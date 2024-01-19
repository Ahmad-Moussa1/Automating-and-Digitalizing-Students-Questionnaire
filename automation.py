import openpyxl as xl
''' values transfer'''
'''
filename = "D:\\Work\\AASTMT\\WORK\\software tasks\\ISO\\August 22\\هندسي\\Questionair Part A\\Questionair Data Reduction\\tmp_PreMarEng_2.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]
mr = ws1.max_row
mc = ws1.max_column
print(mr)
print(mc)
filename1 = "D:\\Work\\AASTMT\\WORK\\software tasks\\ISO\\August 22\\هندسي\\Questionair Part A\\Questionair Data Reduction\\استمارة ملئ بيانات الاستقصاء.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.worksheets[1]
counter = 3
for m in range (2, mr+ 1):
    for n in range(3, mc+1):
        c = ws1.cell(row=m, column=n)
        if c.value == "أوافق تماما":
            c.value = 1
        elif c.value == "أوافق":
            c.value = 2
        elif c.value == "لا أوافق":
            c.value = 3
        ws2.cell(row=counter, column=n).value = c.value
    counter+= 22
wb2.save(str(filename1))
'''
''' part 1 Questions code'''
'''
filename1 = "D:\\Work\\AASTMT\\WORK\\software tasks\\ISO\\August 22\\هندسي\\Questionair Part A\\Questionair Data Reduction\\استمارة ملئ بيانات الاستقصاء.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.worksheets[3]
a = 13
b = 23
m = [4, 5, 6, 7, 8, 9, 26,27,28,29,30,31, 48,49,50,51,52,53, 70,71,72,73,74,75, 92,93,94,95,96,97, 114,115,116,117,118,119, 136,137,138,139,140,141, 158,159,160,161,162,163, 180,181,182,183,184,185]
h = [3, 25, 47, 69, 91, 113, 135, 157, 179]
i = 0
s = 0
while s < 9:
    while a < 64 and b < 74:
        for j, k in zip(range(a, b), range(3, 13)):
            ws2.cell(row=m[i], column=k).value = ws2.cell(row=h[s], column=j).value
        i+=1
        a+=10
        b+=10
    s+=1
    a = 13
    b = 23
wb2.save(str(filename1))
'''

''' Part 2 Questions code'''
'''
filename1 = "D:\\Work\\AASTMT\\WORK\\software tasks\\ISO\\August 22\\هندسي\\Questionair Part A\\Questionair Data Reduction\\استمارة ملئ بيانات الاستقصاء.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.worksheets[3]
a = 74
b = 84
m = [13, 14, 15, 16, 17,  35,36,37,38,39,  57,58,59,60,61,  79,80,81,82,83,  101,102,103,104,105,  123,124,125,126,127,  145,146,147,148,149,  167,168,169,170,171,  189,190,191,192,193 ]
h = [3, 25, 47, 69, 91, 113, 135, 157, 179]
i = 0
s = 0
while s < 9:
    while a < 115 and b < 125:
        for j, k in zip(range(a, b), range(3, 13)):
            ws2.cell(row=m[i], column=k).value = ws2.cell(row=h[s], column=j).value
        i+=1
        a+=10
        b+=10
    s+=1
    a = 74
    b = 84
wb2.save(str(filename1))
'''

''' Part 3 Questions code '''
'''
filename1 = "D:\\Work\\AASTMT\\WORK\\software tasks\\ISO\\August 22\\هندسي\\Questionair Part A\\Questionair Data Reduction\\استمارة ملئ بيانات الاستقصاء.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.worksheets[3]
a = 124
b = 127
m = [20,21,22,  42,43,44,  64,65,66,  86,87,88,  108,109,110,  130,131,132,  152,153,154,  174,175,176,  196,197,198]
h = [3, 25, 47, 69, 91, 113, 135, 157, 179]
i = 0
s = 0
while s < 9:
    while a < 125 and b < 128:
        for j in range(a, b):
            ws2.cell(row=m[i], column=3).value = ws2.cell(row=h[s], column=j).value
            i+=1
        a+=3
        b+=3
    s+=1
    a = 124
    b = 127
wb2.save(str(filename1))
'''

''' Delete transfered values'''
''' delete them manually form the sheet'''
